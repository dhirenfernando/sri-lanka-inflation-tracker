"""Direct, small collectors for the three official DCS headline series."""
from __future__ import annotations

import re
from datetime import date
from html.parser import HTMLParser
from io import BytesIO
from urllib.parse import urljoin
from urllib.request import Request, urlopen

import pdfplumber
from openpyxl import load_workbook

CCPI_PAGE = "https://www.statistics.gov.lk/InflationAndPrices/StaticalInformation/MonthlyCCPI"
NCPI_PAGE = "https://www.statistics.gov.lk/InflationAndPrices/StaticalInformation/MonthlyNCPI"
PPI_PAGE = "https://www.statistics.gov.lk/InflationAndPrices/StaticalInformation/PPI"
MONTHS = {name: number for number, name in enumerate("January February March April May June July August September October November December".split(), 1)}


class DcsError(RuntimeError):
    pass


def _fetch(url: str) -> tuple[bytes, str | None]:
    try:
        with urlopen(Request(url, headers={"User-Agent": "sri-lanka-inflation-tracker/1.0"}), timeout=45) as response:
            data = response.read()
            if not data:
                raise DcsError(f"DCS returned an empty response: {url}")
            return data, response.headers.get("Content-Type")
    except DcsError:
        raise
    except Exception as error:
        raise DcsError(f"Could not fetch DCS source {url}: {error}") from error


class _Links(HTMLParser):
    def __init__(self, section: str | None = None):
        super().__init__(); self.section = section; self.in_section = section is None; self.href = None; self.text = []; self.links = []; self.frames = []
    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "a": self.href, self.text = attrs.get("href"), []
        if tag == "iframe" and attrs.get("src"): self.frames.append(attrs["src"])
    def handle_data(self, data):
        text = " ".join(data.split())
        if self.section:
            lower = text.lower()
            if self.section.lower() in lower: self.in_section = True
            elif "base" in lower and ("ccpi" in lower or "ncpi" in lower): self.in_section = False
        if self.href is not None: self.text.append(data)
    def handle_endtag(self, tag):
        if tag == "a" and self.href is not None:
            self.links.append((" ".join("".join(self.text).split()), self.href, self.in_section)); self.href = None


def _resolve_pdf(page_url: str, label: str, section: str) -> tuple[str, str]:
    page, _ = _fetch(page_url)
    links = _Links(section); links.feed(page.decode("utf-8", errors="replace"))
    matches = [href for text, href, active in links.links if active and text == label]
    if len(matches) != 1:
        raise DcsError(f"{page_url}: expected one {label!r} in {section!r}; found {len(matches)}")
    wrapper = urljoin(page_url, matches[0])
    body, _ = _fetch(wrapper)
    frames = _Links(); frames.feed(body.decode("utf-8", errors="replace"))
    if len(frames.frames) != 1:
        raise DcsError(f"{wrapper}: expected one PDF iframe; found {len(frames.frames)}")
    pdf_url = urljoin(wrapper, frames.frames[0])
    if not pdf_url.split("?", 1)[0].lower().endswith(".pdf"):
        raise DcsError(f"{wrapper}: iframe is not a PDF: {pdf_url}")
    return wrapper, pdf_url


def _pdf(url: str) -> bytes:
    data, content_type = _fetch(url)
    if not data.startswith(b"%PDF-") or (content_type and "pdf" not in content_type.lower()):
        raise DcsError(f"Expected a PDF from {url}; received {content_type!r}")
    return data


def _movement_rows(pdf_bytes: bytes, expected: str) -> list[dict]:
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(page.extract_text(layout=True) or "" for page in pdf.pages)
    if expected not in " ".join(text.upper().split()):
        raise DcsError(f"Downloaded document is not {expected}")
    rows, year = [], None
    pattern = re.compile(r"(?:(20\d{2})\s+)?(" + "|".join(MONTHS) + r")\s+(\d+(?:\.\d+)?(?:\s+-?\d+(?:\.\d+)?){0,3})$")
    for raw in text.splitlines():
        match = pattern.fullmatch(" ".join(raw.split()))
        if not match: continue
        parsed_year, month, values = match.groups(); year = int(parsed_year) if parsed_year else year
        if year is None: raise DcsError(f"Month row has no year: {raw}")
        numbers = [float(value) for value in values.split()]
        rows.append({"period": f"{year:04d}-{MONTHS[month]:02d}-01", "index": numbers[0], "mom": numbers[1] if len(numbers) > 1 else None, "yoy": numbers[2] if len(numbers) > 2 else None, "ma12": numbers[3] if len(numbers) > 3 else None})
    if not rows: raise DcsError("No movement rows found in PDF")
    return rows


def fetch_ccpi() -> tuple[list[dict], str]:
    _, url = _resolve_pdf(CCPI_PAGE, "Movements of the CCPI", "CCPI (Base 2021=100)")
    return _movement_rows(_pdf(url), "MOVEMENTS OF THE CCPI"), url


def fetch_ncpi() -> tuple[list[dict], str]:
    _, url = _resolve_pdf(NCPI_PAGE, "Movements of the NCPI", "NCPI (Base 2021=100)")
    return _movement_rows(_pdf(url), "MOVEMENTS OF THE NCPI"), url


def _ppi_period(value, number_format: str) -> date:
    # DCS encodes some headers as a date whose *day* is the two-digit year.
    if number_format == "d-mmm" and isinstance(value, date) and 14 <= value.day <= 99:
        return date(2000 + value.day, value.month, 1)
    if isinstance(value, date): return date(value.year, value.month, 1)
    match = re.fullmatch(r"([A-Za-z]+)-(\d{2})(?:\*{1,2})?", str(value))
    if match:
        month, year = match.groups()
        return date(2000 + int(year), {m[:3]: n for n, m in enumerate("Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), 1)}[month[:3]], 1)
    raise DcsError(f"Unsupported PPI period header {value!r} ({number_format!r})")


def fetch_ppi() -> tuple[list[dict], str]:
    page, _ = _fetch(PPI_PAGE); parser = _Links(); parser.feed(page.decode("utf-8", errors="replace"))
    links = [href for text, href, _ in parser.links if text == "Movements of PPI"]
    if len(links) != 1: raise DcsError(f"{PPI_PAGE}: expected one 'Movements of PPI'; found {len(links)}")
    url = urljoin(PPI_PAGE, links[0]); data, _ = _fetch(url)
    if not data.startswith(b"PK\x03\x04"): raise DcsError(f"Expected an XLSX workbook from {url}")
    return parse_ppi_workbook(data), url


def parse_ppi_workbook(data: bytes) -> list[dict]:
    try: sheet = load_workbook(BytesIO(data), data_only=True)["Sheet1"]
    except Exception as error: raise DcsError(f"Could not open PPI workbook: {error}") from error
    if "PRODUCER PRICE INDEX" not in str(sheet["B1"].value).upper() or sheet["B4"].value != "Producer Price Index (PPI)":
        raise DcsError("PPI workbook identity check failed")
    row = next((r for r in range(1, sheet.max_row + 1) if sheet.cell(r, 2).value == "Producer Price Index (PPI)"), None)
    if row is None: raise DcsError("PPI aggregate row not found")
    headers = [column for column in range(3, sheet.max_column + 1) if sheet.cell(3, column).value is not None]
    periods = []
    for position, column in enumerate(headers):
        header = sheet.cell(3, column)
        if header.value == "July":
            if position == 0 or position == len(headers) - 1:
                raise DcsError("PPI unyear-ed July header is at an edge")
            before = _ppi_period(sheet.cell(3, headers[position - 1]).value, sheet.cell(3, headers[position - 1]).number_format)
            after = _ppi_period(sheet.cell(3, headers[position + 1]).value, sheet.cell(3, headers[position + 1]).number_format)
            if (before, after) != (date(2024, 6, 1), date(2024, 8, 1)):
                raise DcsError(f"PPI July header has ambiguous neighbours: {before}, {after}")
            periods.append((column, date(2024, 7, 1)))
        else:
            periods.append((column, _ppi_period(header.value, header.number_format)))
    rows = []
    for column, period in periods:
        value = sheet.cell(row, column).value
        if not isinstance(value, (int, float)): raise DcsError(f"PPI aggregate {period:%Y-%m} is not numeric")
        rows.append({"period": period.isoformat(), "index": float(value)})
    if not rows: raise DcsError("No PPI aggregate observations found")
    return rows
